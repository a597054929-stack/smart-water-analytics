#!/usr/bin/env python3
"""
按建筑物聚合预测 - 路氹城區 Top20 建筑
输出: public/data/predictions_by_building.json
"""

import json
import numpy as np
from datetime import datetime, timedelta
from sklearn.linear_model import LinearRegression
from collections import defaultdict
import os

# 配置
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'output', 'all_data.json')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'output', 'predictions_by_building.json')
PREDICT_DAYS = 7
TOP_N_BUILDINGS = 20
TARGET_DMA = '路氹城區'  # 目标DMA


def load_data():
    """加载数据"""
    with open(DATA_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_full_building_data(building_meters, dates):
    """从 Excel 加载建筑物的完整数据（聚合多个水表）"""
    import os
    import openpyxl
    
    cons_dir = os.environ.get('CONS_DIR', os.path.join(SCRIPT_DIR, '..', 'data', 'input', 'consumption'))
    
    # 初始化建筑物时间序列
    building_series = {b: [0] * len(dates) for b in building_meters.keys()}
    
    print(f"Loading full building data from {len(dates)} days...")
    
    # 收集所有需要追踪的水表
    all_meters = set()
    for meters in building_meters.values():
        all_meters.update(meters)
    
    print(f"  Tracking {len(all_meters)} meters across {len(building_meters)} buildings")
    
    for i, date in enumerate(dates):
        excel_file = os.path.join(cons_dir, f'{date}.xlsx')
        if not os.path.exists(excel_file):
            continue
        
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            
            # 找到水表编号列和小时列
            meter_col = None
            hour_cols = []
            for j, h in enumerate(headers):
                if h and '錶位編號' in str(h):
                    meter_col = j
                elif h and ':' in str(h):
                    hour_cols.append(j)
            
            if meter_col is None or not hour_cols:
                wb.close()
                continue
            
            # 读取数据并按建筑物聚合
            meter_usage = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                mid = str(row[meter_col] or '').strip()
                if mid in all_meters:
                    total = sum(float(row[h] or 0) for h in hour_cols) / 1000
                    if 0 < total < 5000:
                        meter_usage[mid] = total
            
            # 按建筑物聚合
            for building, meters in building_meters.items():
                day_total = sum(meter_usage.get(m, 0) for m in meters)
                if day_total > 0:
                    building_series[building][i] = day_total
            
            wb.close()
        except Exception as e:
            print(f"  Error reading {date}: {e}")
            continue
        
        if (i + 1) % 20 == 0:
            print(f"  Processed {i + 1}/{len(dates)} days...")
    
    print(f"  Done!")
    return building_series


def get_cotai_dma_key(data):
    """获取路氹城區的编码key"""
    for day in data.get('dma', []):
        for dma_key in day.get('dmas', {}).keys():
            if '路' in dma_key and '城' in dma_key:
                return dma_key
            # 检查特殊编码
            if len(dma_key) > 1 and dma_key[1] and ord(dma_key[1]) == 0xebf3:
                return dma_key
    return None


def aggregate_by_building(data, cotai_key):
    """按建筑物聚合用水量"""
    building_daily = defaultdict(lambda: defaultdict(float))
    building_info = {}
    
    # 从 Top20 DMA 数据中提取路氹城區的水表
    for day_data in data.get('top20dma', []):
        date = day_data['date']
        by_dma = day_data.get('byDma', {})
        
        # 找到路氹城區的数据
        cotai_data = None
        for dma_key, meters in by_dma.items():
            if dma_key == cotai_key or ('路' in dma_key and '城' in dma_key):
                cotai_data = meters
                break
        
        if not cotai_data:
            continue
        
        # 按建筑物聚合
        for meter in cotai_data:
            building = meter.get('buildingName', '未知')
            if not building:
                building = meter.get('contractId', '未知')
            
            building_daily[building][date] += meter['total']
            
            if building not in building_info:
                building_info[building] = {
                    'dma': cotai_key,
                    'meters': set(),
                    'propertyType': meter.get('propertyType', '')
                }
            building_info[building]['meters'].add(meter['meterId'])
    
    return building_daily, building_info


def get_top_buildings(building_daily, n=20):
    """获取总用水量最高的N个建筑物"""
    building_totals = {}
    for building, daily in building_daily.items():
        building_totals[building] = sum(daily.values())
    
    top_buildings = sorted(building_totals.items(), key=lambda x: x[1], reverse=True)[:n]
    return top_buildings


def create_features(date_str, values, idx):
    """创建特征"""
    # 确保 date_str 是字符串
    if isinstance(date_str, list):
        date_str = date_str[0] if date_str else ''
    date = datetime.strptime(str(date_str), '%Y-%m-%d')
    features = [
        date.weekday(),  # 星期几
        date.month,  # 月份
        values[idx - 1] if idx >= 1 else values[idx],  # 前1天
        np.mean(values[max(0, idx - 7):idx]) if idx >= 1 else values[idx],  # 前7天平均
        np.mean(values[max(0, idx - 14):idx]) if idx >= 2 else values[idx],  # 前14天平均
        idx  # 趋势
    ]
    return features


def predict_building(building_name, daily_data, dates, predict_days=7):
    """预测单个建筑物"""
    # 构建完整时间序列
    if isinstance(daily_data, list):
        values = daily_data
    else:
        values = [daily_data.get(d, 0) for d in dates]
    n = len(values)
    
    if n < 14:  # 至少需要14天数据
        return None
    
    # 准备训练数据
    X_train = []
    y_train = []
    
    for i in range(14, n):
        features = create_features(dates, values, i)
        X_train.append(features)
        y_train.append(values[i])
    
    if len(X_train) < 10:
        return None
    
    # 训练模型
    model = LinearRegression()
    model.fit(np.array(X_train), np.array(y_train))
    score = model.score(np.array(X_train), np.array(y_train))
    
    # 计算所有历史数据的拟合值
    fitted_values = []
    for i in range(14, n):  # 从第15天开始
        features = create_features(dates, values, i)
        fitted = model.predict([features])[0]
        fitted_values.append({
            'date': dates[i],
            'actual': round(values[i], 2),
            'fitted': round(max(0, fitted), 2)
        })
    
    # 预测未来7天
    last_date = datetime.strptime(dates[-1], '%Y-%m-%d')
    predictions = []
    
    recent_values = list(values[-14:])
    current_values = list(values)
    
    for day_offset in range(1, predict_days + 1):
        pred_date = last_date + timedelta(days=day_offset)
        pred_idx = n + day_offset - 1
        
        features = [
            pred_date.weekday(),
            pred_date.month,
            current_values[-1],
            np.mean(recent_values[-7:]),
            np.mean(recent_values[-14:]),
            pred_idx
        ]
        
        pred = model.predict([features])[0]
        pred = max(0, pred)
        
        predictions.append({
            'date': pred_date.strftime('%Y-%m-%d'),
            'value': round(pred, 2)
        })
        
        current_values.append(pred)
        recent_values.append(pred)
    
    return {
        'building': building_name,
        'fitted': fitted_values,
        'predictions': predictions,
        'modelScore': round(score, 4),
        'avgHistorical': round(np.mean([v for v in values if v > 0]), 2),
        'trend': 'up' if model.coef_[-1] > 0 else 'down'
    }


def main():
    print("Loading data...")
    data = load_data()
    
    print("Finding Cotai DMA key...")
    cotai_key = get_cotai_dma_key(data)
    if not cotai_key:
        print("ERROR: Cannot find Cotai DMA")
        return
    print(f"   Found Cotai DMA key")
    
    print("Aggregating by building...")
    building_daily, building_info = aggregate_by_building(data, cotai_key)
    print(f"   Found {len(building_daily)} buildings")
    
    print(f"Getting top {TOP_N_BUILDINGS} buildings...")
    top_buildings = get_top_buildings(building_daily, TOP_N_BUILDINGS)
    
    dates = data.get('dates', [])
    print(f"\nPredicting next {PREDICT_DAYS} days for {len(top_buildings)} buildings...")
    
    results = []
    for building_name, total in top_buildings:
        daily_data = building_daily.get(building_name, {})
        info = building_info.get(building_name, {})
        
        result = predict_building(building_name, daily_data, dates, PREDICT_DAYS)
        
        if result:
            result['totalHistorical'] = round(total, 2)
            result['meterCount'] = len(info.get('meters', set()))
            result['propertyType'] = info.get('propertyType', '')
            results.append(result)
            non_zero_days = sum(1 for v in daily_data.values() if v > 0)
            print(f"   [OK] {building_name}: score={result['modelScore']}, trend={result['trend']}, days={non_zero_days}")
        else:
            print(f"   [SKIP] {building_name}: insufficient data")
    
    # 保存结果
    output = {
        'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'dma': cotai_key,
        'historicalRange': {
            'start': dates[0],
            'end': dates[-1],
            'days': len(dates)
        },
        'predictionDays': PREDICT_DAYS,
        'totalBuildings': len(results),
        'predictions': results
    }
    
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\nDone! Saved {len(results)} building predictions to {OUTPUT_PATH}")
    print(f"Historical data: {dates[0]} to {dates[-1]} ({len(dates)} days)")
    print(f"Prediction period: {PREDICT_DAYS} days")


if __name__ == '__main__':
    main()

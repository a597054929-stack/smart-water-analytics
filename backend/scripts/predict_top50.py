#!/usr/bin/env python3
"""
预测 Top50 水表未来7天用水量 (线性回归)
输出: public/data/predictions.json
"""

import json
import numpy as np
from datetime import datetime, timedelta
from sklearn.linear_model import LinearRegression
from collections import defaultdict
import os

# 配置
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'output', 'all_data.json')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, '..', 'data', 'output', 'predictions.json')
PREDICT_DAYS = 7
TOP_N = 50


def load_data():
    """加载数据"""
    with open(DATA_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_full_meter_data(meter_ids, dates):
    """从 Excel 加载完整水表数据"""
    import os
    import openpyxl
    
    cons_dir = os.environ.get('CONS_DIR', os.path.join(SCRIPT_DIR, '..', 'data', 'input', 'consumption'))
    
    # 初始化时间序列
    meter_series = {mid: [0] * len(dates) for mid in meter_ids}
    
    print(f"Loading full data for {len(meter_ids)} meters from {len(dates)} days...")
    
    for i, date in enumerate(dates):
        excel_file = os.path.join(cons_dir, f'{date}.xlsx')
        if not os.path.exists(excel_file):
            continue
        
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            
            # 找到水表编号列和小时列
            meter_col = None
            hour_cols = []
            for j, h in enumerate(headers):
                if h and '錶位編號' in str(h):
                    meter_col = j
                elif h and ':' in str(h):
                    hour_cols.append(j)
            
            if meter_col is None or not hour_cols:
                wb.close()
                continue
            
            # 读取数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                mid = str(row[meter_col] or '').strip()
                if mid in meter_ids:
                    total = sum(float(row[h] or 0) for h in hour_cols) / 1000  # 升转立方米
                    if 0 < total < 5000:  # 过滤异常值
                        meter_series[mid][i] = total
            
            wb.close()
        except Exception as e:
            print(f"  Error reading {date}: {e}")
            continue
        
        if (i + 1) % 20 == 0:
            print(f"  Processed {i + 1}/{len(dates)} days...")
    
    print(f"  Done!")
    return meter_series


def get_top_meters(data, n=50):
    """获取 Top N 水表（按总用水量排序）"""
    meter_totals = defaultdict(float)
    meter_info = {}
    
    for day_data in data.get('top20', []):
        for meter in day_data.get('top20', []):
            mid = meter['meterId']
            meter_totals[mid] += meter['total']
            if mid not in meter_info:
                meter_info[mid] = {
                    'dma': meter.get('dma', ''),
                    'propertyType': meter.get('propertyType', ''),
                    'buildingName': meter.get('buildingName', '')
                }
    
    top_meters = sorted(meter_totals.items(), key=lambda x: x[1], reverse=True)[:n]
    return top_meters, meter_info


def build_meter_timeseries(data):
    """构建每个水表的完整时间序列"""
    dates = data.get('dates', [])
    n_days = len(dates)
    
    # 初始化时间序列
    meter_series = defaultdict(lambda: [0] * n_days)
    
    # 从 top20 数据填充
    for i, day_data in enumerate(data.get('top20', [])):
        for meter in day_data.get('top20', []):
            mid = meter['meterId']
            meter_series[mid][i] = meter['total']
    
    return dict(meter_series), dates


def create_features(dates, values, idx):
    """
    为预测创建特征
    - 星期几 (0-6)
    - 月份 (1-12)
    - 前1天用水量
    - 前7天平均
    - 趋势 (天数索引)
    """
    date = datetime.strptime(dates[idx], '%Y-%m-%d')
    
    features = [
        date.weekday(),  # 星期几
        date.month,  # 月份
        values[idx - 1] if idx >= 1 else values[idx],  # 前1天
        np.mean(values[max(0, idx - 7):idx]) if idx >= 1 else values[idx],  # 前7天平均
        idx  # 趋势
    ]
    return features


def predict_meter(meter_id, values, dates, predict_days=7):
    """使用线性回归预测单个水表"""
    n = len(values)
    
    if n < 14:  # 至少需要14天数据
        return None
    
    # 准备训练数据
    X_train = []
    y_train = []
    
    for i in range(7, n):  # 从第8天开始（需要前7天数据）
        features = create_features(dates, values, i)
        X_train.append(features)
        y_train.append(values[i])
    
    if len(X_train) < 10:
        return None
    
    # 训练模型
    model = LinearRegression()
    model.fit(np.array(X_train), np.array(y_train))
    
    # 计算所有历史数据的拟合值（用于验证模型和展示）
    fitted_values = []
    for i in range(7, n):  # 从第8天开始（需要前7天数据）
        features = create_features(dates, values, i)
        fitted = model.predict([features])[0]
        fitted_values.append({
            'date': dates[i],
            'actual': round(values[i], 2),
            'fitted': round(max(0, fitted), 2)
        })
    
    # 预测未来7天
    last_date = datetime.strptime(dates[-1], '%Y-%m-%d')
    predictions = []
    
    # 使用最后7天的值作为初始滞后特征
    recent_values = list(values[-7:])
    current_values = list(values)
    
    for day_offset in range(1, predict_days + 1):
        pred_date = last_date + timedelta(days=day_offset)
        pred_idx = n + day_offset - 1
        
        # 构建特征
        features = [
            pred_date.weekday(),
            pred_date.month,
            current_values[-1],  # 前1天
            np.mean(recent_values[-7:]),  # 前7天平均
            pred_idx
        ]
        
        pred = model.predict([features])[0]
        pred = max(0, pred)  # 用水量不能为负
        
        predictions.append({
            'date': pred_date.strftime('%Y-%m-%d'),
            'value': round(pred, 2)
        })
        
        # 更新滞后值
        current_values.append(pred)
        recent_values.append(pred)
    
    return {
        'meterId': meter_id,
        'fitted': fitted_values,  # 最近7天的拟合值
        'predictions': predictions,
        'modelScore': round(model.score(np.array(X_train), np.array(y_train)), 4),
        'avgHistorical': round(np.mean(values), 2),
        'trend': 'up' if model.coef_[-1] > 0 else 'down'
    }


def main():
    print("Loading data...")
    data = load_data()
    
    print("Getting top 50 meters...")
    top_meters, meter_info = get_top_meters(data, TOP_N)
    
    print("Building time series from Top20 data...")
    meter_series, dates = build_meter_timeseries(data)
    
    # 检查哪些水表数据不足，从 Excel 加载完整数据
    sparse_meters = []
    for meter_id, _ in top_meters:
        values = meter_series.get(meter_id, [])
        non_zero = sum(1 for v in values if v > 0)
        if non_zero < 30:  # 少于30天有数据
            sparse_meters.append(meter_id)
    
    if sparse_meters:
        print(f"\n{len(sparse_meters)} meters have sparse data, loading from Excel...")
        full_data = load_full_meter_data(set(sparse_meters), dates)
        for mid in sparse_meters:
            if mid in full_data:
                meter_series[mid] = full_data[mid]
    
    print(f"\nPredicting next {PREDICT_DAYS} days for {len(top_meters)} meters...")
    results = []
    
    for meter_id, total in top_meters:
        values = meter_series.get(meter_id, [])
        
        if not values:
            print(f"  Skipping {meter_id}: no data")
            continue
        
        result = predict_meter(meter_id, values, dates, PREDICT_DAYS)
        
        if result:
            result['totalHistorical'] = round(total, 2)
            result['info'] = meter_info.get(meter_id, {})
            results.append(result)
            non_zero_days = sum(1 for v in values if v > 0)
            print(f"  [OK] {meter_id}: score={result['modelScore']}, trend={result['trend']}, days={non_zero_days}")
        else:
            print(f"  [SKIP] {meter_id}: insufficient data")
    
    # 保存结果
    output = {
        'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'historicalRange': {
            'start': dates[0],
            'end': dates[-1],
            'days': len(dates)
        },
        'predictionDays': PREDICT_DAYS,
        'totalMeters': len(results),
        'predictions': results
    }
    
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\nDone! Saved {len(results)} predictions to {OUTPUT_PATH}")
    print(f"Historical data: {dates[0]} to {dates[-1]} ({len(dates)} days)")
    print(f"Prediction period: {PREDICT_DAYS} days")


if __name__ == '__main__':
    main()
